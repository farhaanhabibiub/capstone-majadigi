import csv, random

random.seed(99)

vehicles = [
    # jenis, model, merk, tipe, njkb_2024_base, tahun_start, tahun_end, depr_rate
    # Sepeda Motor
    ('Sepeda Motor','Matic','Honda','125cc',        23_000_000, 2015, 2024, 0.07),
    ('Sepeda Motor','Matic','Honda','160cc',        38_500_000, 2021, 2024, 0.07),
    ('Sepeda Motor','Bebek','Honda','110cc',        18_500_000, 2015, 2024, 0.08),
    ('Sepeda Motor','Sport','Honda','150cc',        32_000_000, 2018, 2024, 0.07),
    ('Sepeda Motor','Trail / Adventure','Honda','150cc', 39_500_000, 2018, 2024, 0.07),
    ('Sepeda Motor','Matic','Yamaha','155cc',       35_000_000, 2018, 2024, 0.07),
    ('Sepeda Motor','Sport','Yamaha','155cc',       44_000_000, 2017, 2024, 0.07),
    ('Sepeda Motor','Naked Sport','Yamaha','155cc', 34_500_000, 2019, 2024, 0.07),
    ('Sepeda Motor','Bebek','Yamaha','110cc',       15_500_000, 2015, 2023, 0.08),
    ('Sepeda Motor','Sport','Kawasaki','250cc',     72_000_000, 2019, 2024, 0.06),
    ('Sepeda Motor','Trail / Adventure','Kawasaki','150cc', 41_000_000, 2019, 2024, 0.07),
    ('Sepeda Motor','Naked Sport','Kawasaki','250cc', 75_000_000, 2020, 2024, 0.06),
    ('Sepeda Motor','Matic','Suzuki','125cc',       22_500_000, 2018, 2024, 0.07),
    ('Sepeda Motor','Sport','TVS','150cc',          25_000_000, 2020, 2024, 0.08),
    ('Sepeda Motor','Naked Sport','Royal Enfield','300cc ke atas', 82_000_000, 2021, 2024, 0.06),
    # Mobil Penumpang
    ('Mobil Penumpang','City Car','Toyota','1.0L',    185_000_000, 2015, 2024, 0.06),
    ('Mobil Penumpang','City Car','Honda','1.2L',     195_000_000, 2018, 2024, 0.06),
    ('Mobil Penumpang','City Car','Daihatsu','1.0L',  170_000_000, 2015, 2024, 0.06),
    ('Mobil Penumpang','City Car','Kia','1.2L',       235_000_000, 2020, 2024, 0.06),
    ('Mobil Penumpang','Hatchback','Toyota','1.5L',   295_000_000, 2018, 2024, 0.05),
    ('Mobil Penumpang','Hatchback','Honda','1.5L',    285_000_000, 2018, 2022, 0.06),
    ('Mobil Penumpang','Hatchback','Suzuki','1.5L',   278_000_000, 2020, 2024, 0.06),
    ('Mobil Penumpang','Hatchback','Hyundai','1.5L',  258_000_000, 2021, 2024, 0.06),
    ('Mobil Penumpang','Hatchback','Kia','1.2L',      245_000_000, 2021, 2024, 0.06),
    ('Mobil Penumpang','Sedan','Honda','1.5L',        610_000_000, 2019, 2024, 0.05),
    ('Mobil Penumpang','Sedan','Toyota','2.4L',       980_000_000, 2019, 2024, 0.04),
    ('Mobil Penumpang','Sedan','Hyundai','1.5L',      348_000_000, 2022, 2024, 0.05),
    # Mobil Beban
    ('Mobil Beban / Pick Up','Pick Up','Daihatsu','1.3L',   215_000_000, 2015, 2024, 0.06),
    ('Mobil Beban / Pick Up','Pick Up','Isuzu','1.5L',      245_000_000, 2018, 2024, 0.06),
    ('Mobil Beban / Pick Up','Pick Up','Mitsubishi','1.5L', 238_000_000, 2018, 2024, 0.06),
    ('Mobil Beban / Pick Up','Double Cabin','Toyota','2.4L',    490_000_000, 2018, 2024, 0.05),
    ('Mobil Beban / Pick Up','Double Cabin','Mitsubishi','2.4L',455_000_000, 2019, 2024, 0.05),
    ('Mobil Beban / Pick Up','Double Cabin','Isuzu','3.0L Diesel',   515_000_000, 2020, 2024, 0.05),
    ('Mobil Beban / Pick Up','Double Cabin','Nissan','2.5L Diesel',  428_000_000, 2018, 2022, 0.05),
    ('Mobil Beban / Pick Up','Flat Deck','Isuzu','3.0L Diesel',      485_000_000, 2018, 2024, 0.05),
    # MPV
    ('Minibus / MPV','Low MPV','Toyota','1.3L',      265_000_000, 2015, 2024, 0.05),
    ('Minibus / MPV','Low MPV','Daihatsu','1.3L',    248_000_000, 2015, 2024, 0.05),
    ('Minibus / MPV','Low MPV','Suzuki','1.5L',      242_000_000, 2018, 2024, 0.05),
    ('Minibus / MPV','Low MPV','Nissan','1.5L',      228_000_000, 2018, 2021, 0.06),
    ('Minibus / MPV','Mid MPV','Toyota','2.0L',      405_000_000, 2018, 2024, 0.05),
    ('Minibus / MPV','Mid MPV','Toyota','2.4L',      460_000_000, 2021, 2024, 0.04),
    ('Minibus / MPV','Mid MPV','Mitsubishi','1.5L',  320_000_000, 2018, 2024, 0.05),
    ('Minibus / MPV','High MPV','Toyota','2.4L',   1_350_000_000, 2018, 2024, 0.04),
    ('Minibus / MPV','High MPV','Honda','2.4L',      895_000_000, 2018, 2022, 0.04),
    # SUV
    ('SUV / Jeep','Compact SUV','Toyota','1.5L',     295_000_000, 2018, 2024, 0.05),
    ('SUV / Jeep','Compact SUV','Daihatsu','1.5L',   275_000_000, 2018, 2024, 0.05),
    ('SUV / Jeep','Compact SUV','Honda','1.5L',      348_000_000, 2022, 2024, 0.05),
    ('SUV / Jeep','Compact SUV','Hyundai','1.5L',    395_000_000, 2020, 2024, 0.05),
    ('SUV / Jeep','Compact SUV','Kia','1.5L',        468_000_000, 2021, 2024, 0.05),
    ('SUV / Jeep','Mid SUV','Toyota','2.4L',         620_000_000, 2018, 2024, 0.04),
    ('SUV / Jeep','Mid SUV','Honda','1.5L',          595_000_000, 2020, 2024, 0.05),
    ('SUV / Jeep','Mid SUV','Mitsubishi','2.5L Diesel', 640_000_000, 2019, 2024, 0.04),
    ('SUV / Jeep','Mid SUV','Ford','2.0L',           980_000_000, 2022, 2024, 0.04),
    ('SUV / Jeep','Mid SUV','Chevrolet','1.5L',      548_000_000, 2019, 2021, 0.05),
    ('SUV / Jeep','Large SUV','Toyota','3.5L',     1_850_000_000, 2021, 2024, 0.04),
    ('SUV / Jeep','Jeep Off-Road','Jeep','2.0L',   1_380_000_000, 2021, 2024, 0.04),
    # Bus
    ('Bus','Micro Bus','Isuzu','4 Silinder Diesel',     720_000_000, 2018, 2024, 0.05),
    ('Bus','Micro Bus','Mitsubishi','4 Silinder Diesel', 695_000_000, 2018, 2024, 0.05),
    ('Bus','Medium Bus','Hino','4 Silinder Diesel',    1_050_000_000, 2018, 2024, 0.04),
    ('Bus','Medium Bus','Mitsubishi','4 Silinder Diesel', 980_000_000, 2018, 2024, 0.04),
    ('Bus','Bus Besar','Hino','6 Silinder Diesel',     2_350_000_000, 2018, 2024, 0.04),
    ('Bus','Bus Besar','Scania','6 Silinder Diesel',   3_800_000_000, 2020, 2024, 0.03),
    ('Bus','Bus Besar','Mercedes-Benz','Turbodiesel CNG', 3_600_000_000, 2020, 2024, 0.03),
    ('Bus','Bus Besar','MAN','6 Silinder Diesel',      3_200_000_000, 2020, 2024, 0.03),
    # Truk
    ('Truk','Truk Ringan','Mitsubishi','4 Silinder Diesel', 355_000_000, 2018, 2024, 0.05),
    ('Truk','Truk Ringan','Isuzu','4 Silinder Diesel',      380_000_000, 2018, 2024, 0.05),
    ('Truk','Truk Sedang','Hino','6 Silinder Diesel',       830_000_000, 2018, 2024, 0.04),
    ('Truk','Truk Sedang','Isuzu','6 Silinder Diesel',      810_000_000, 2018, 2024, 0.04),
    ('Truk','Truk Berat','Isuzu','Turbodiesel',           1_250_000_000, 2018, 2024, 0.04),
    ('Truk','Truk Berat','Hino','Turbodiesel',            1_300_000_000, 2018, 2024, 0.04),
    ('Truk','Truk Berat','Mercedes-Benz','Turbodiesel',   2_800_000_000, 2020, 2024, 0.03),
    ('Truk','Truk Gandeng','Scania','Intercooler Diesel', 4_500_000_000, 2020, 2024, 0.03),
    ('Truk','Truk Gandeng','Volvo','Intercooler Diesel',  4_800_000_000, 2020, 2024, 0.03),
    # Kendaraan Khusus
    ('Kendaraan Khusus','Ambulans','Toyota','4 Silinder Diesel',     470_000_000, 2018, 2024, 0.05),
    ('Kendaraan Khusus','Ambulans','Mitsubishi','4 Silinder Diesel', 445_000_000, 2018, 2024, 0.05),
    ('Kendaraan Khusus','Ambulans','Isuzu','4 Silinder Diesel',      480_000_000, 2019, 2024, 0.05),
    ('Kendaraan Khusus','Pemadam Kebakaran','Isuzu','6 Silinder Diesel',     1_700_000_000, 2018, 2024, 0.04),
    ('Kendaraan Khusus','Pemadam Kebakaran','Mercedes-Benz','6 Silinder Diesel', 3_200_000_000, 2019, 2024, 0.03),
    ('Kendaraan Khusus','Derek','Toyota','4 Silinder Diesel',        380_000_000, 2018, 2024, 0.05),
    ('Kendaraan Khusus','Mixer Beton','Isuzu','6 Silinder Diesel',   1_200_000_000, 2018, 2024, 0.04),
    ('Kendaraan Khusus','Mixer Beton','Mercedes-Benz','6 Silinder Diesel', 2_100_000_000, 2019, 2024, 0.04),
]

rows = []
for jenis, model, merk, tipe, base_2024, y_start, y_end, depr in vehicles:
    for tahun in range(y_start, y_end + 1):
        age = 2024 - tahun
        factor = max(0.30, 1.0 - depr * age)
        noise = 1 + random.uniform(-0.03, 0.03)
        njkb = round(base_2024 * factor * noise / 1_000_000) * 1_000_000
        rows.append([jenis, model, merk, tipe, str(tahun), str(njkb)])

csv_path = r'C:\Users\farhaanhabibi\AndroidStudioProjects\capstonemajadigi\assets\data\njkb_database.csv'
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['jenis','model','merk','tipe','tahun','njkb'])
    w.writerows(rows)

print(f"Generated {len(rows)} records to {csv_path}")

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Database NJKB'

blue_fill = PatternFill(start_color='0065FF', end_color='0065FF', fill_type='solid')
white_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
hdr_align = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['Jenis Kendaraan','Model','Merk','Tipe','Tahun','NJKB (Rp)']
ws.row_dimensions[1].height = 28
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = white_font; c.fill = blue_fill; c.alignment = hdr_align; c.border = border

alt_fill = PatternFill(start_color='EBF3FF', end_color='EBF3FF', fill_type='solid')
for ri, row in enumerate(rows, 2):
    ws.row_dimensions[ri].height = 17
    fill = alt_fill if ri % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci)
        c.font = Font(name='Calibri', size=10)
        c.border = border
        if fill: c.fill = fill
        if ci == 6:
            c.value = int(val)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.number_format = '#,##0'
        else:
            c.value = val
            c.alignment = Alignment(horizontal='left', vertical='center')

for ci, w in zip(range(1,7), [25,18,16,22,8,18]):
    ws.column_dimensions[ws.cell(1,ci).column_letter].width = w
ws.freeze_panes = 'A2'

xlsx_path = r'C:\Users\farhaanhabibi\AndroidStudioProjects\capstonemajadigi\assets\data\njkb_database.xlsx'
wb.save(xlsx_path)
print(f"Excel saved: {xlsx_path}")
